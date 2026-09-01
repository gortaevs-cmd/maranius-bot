#!/usr/bin/env python3
"""Prepare or apply the September 2026 legacy-contact migration.

The manifest contains only Telegram IDs and is deliberately not stored in Git.
Create it locally from the two supplied XLSX exports, inspect the dry run, then
stage it as ``legacy_migration_pending.json`` in the production runtime
directory.  The deployed bot consumes that file before it starts polling, so
the migration cannot race with normal user updates.  It never copies historical
consent or marketing opt-ins.
"""

from __future__ import annotations

import argparse
import json
import os
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

PROJECT_DIR = Path(__file__).resolve().parents[1]
if str(PROJECT_DIR) not in sys.path:
    sys.path.insert(0, str(PROJECT_DIR))

from integrations import legacy_contacts, user_registry  # noqa: E402
from integrations.json_storage import load_json, save_json  # noqa: E402

MANIFEST_VERSION = 1


def _now_iso() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def _read_xlsx_rows(path: Path) -> list[tuple[Any, ...]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "Для подготовки манифеста из XLSX нужен openpyxl. "
            "Подготовьте манифест на рабочем Mac, а в production передавайте уже JSON."
        ) from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return list(workbook.active.iter_rows(values_only=True))
    finally:
        workbook.close()


def _header_index(headers: Iterable[Any], expected: str, *, file_label: str) -> int:
    normalized = [str(value or "").strip() for value in headers]
    try:
        return normalized.index(expected)
    except ValueError as exc:
        raise ValueError(f"{file_label}: required column {expected!r} was not found") from exc


def build_manifest(contacts_xlsx: Path, recipients_xlsx: Path) -> dict[str, Any]:
    contacts = _read_xlsx_rows(contacts_xlsx)
    recipients = _read_xlsx_rows(recipients_xlsx)
    if len(contacts) < 2 or len(recipients) < 2:
        raise ValueError("Both XLSX files must contain a header and at least one data row")

    contacts_id_col = _header_index(
        contacts[0], "ID пользователя", file_label=contacts_xlsx.name
    )
    vip_col = _header_index(contacts[0], "%VIP_клиент%", file_label=contacts_xlsx.name)
    recipients_id_col = _header_index(
        recipients[0], "ID пользователя", file_label=recipients_xlsx.name
    )

    contact_vip: dict[int, bool] = {}
    for row_number, row in enumerate(contacts[1:], start=2):
        if max(contacts_id_col, vip_col) >= len(row):
            raise ValueError(f"{contacts_xlsx.name}: incomplete row {row_number}")
        user_id = legacy_contacts.normalize_user_id(row[contacts_id_col])
        if user_id in contact_vip:
            raise ValueError(f"{contacts_xlsx.name}: duplicate Telegram ID at row {row_number}")
        vip_raw = str(row[vip_col] or "").strip().casefold()
        if vip_raw not in {"true", "false"}:
            raise ValueError(f"{contacts_xlsx.name}: invalid VIP value at row {row_number}")
        contact_vip[user_id] = vip_raw == "true"

    recipient_ids: set[int] = set()
    for row_number, row in enumerate(recipients[1:], start=2):
        if recipients_id_col >= len(row):
            raise ValueError(f"{recipients_xlsx.name}: incomplete row {row_number}")
        user_id = legacy_contacts.normalize_user_id(row[recipients_id_col])
        if user_id in recipient_ids:
            raise ValueError(f"{recipients_xlsx.name}: duplicate Telegram ID at row {row_number}")
        recipient_ids.add(user_id)

    unknown_recipients = recipient_ids.difference(contact_vip)
    if unknown_recipients:
        raise ValueError(
            f"{recipients_xlsx.name}: {len(unknown_recipients)} recipient IDs are absent from contacts"
        )

    vip_ids = {user_id for user_id, is_vip in contact_vip.items() if is_vip}
    active_ids = recipient_ids | vip_ids
    inactive_ids = set(contact_vip).difference(active_ids)
    return {
        "schema_version": MANIFEST_VERSION,
        "source": "legacy_broadcast_delivery_2026-09-01",
        "prepared_at": _now_iso(),
        "active_user_ids": sorted(active_ids),
        "vip_user_ids": sorted(vip_ids),
        "inactive_user_ids": sorted(inactive_ids),
    }


def load_manifest(path: Path) -> dict[str, Any]:
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise ValueError(f"Cannot read migration manifest {path}") from exc
    if not isinstance(data, dict) or data.get("schema_version") != MANIFEST_VERSION:
        raise ValueError("Unsupported or malformed migration manifest")
    required = ("source", "active_user_ids", "vip_user_ids", "inactive_user_ids")
    if any(key not in data for key in required):
        raise ValueError("Migration manifest is missing required fields")
    return data


def _print_summary(
    *, manifest: dict[str, Any], result: dict[str, int], applied: bool
) -> None:
    state = "applied" if applied else "preview"
    print(f"Legacy migration {state}: {manifest['source']}")
    print(
        "Input: "
        f"active={len(manifest['active_user_ids'])}, "
        f"vip={len(manifest['vip_user_ids'])}, "
        f"inactive={len(manifest['inactive_user_ids'])}"
    )
    print(
        "Result: "
        f"active_created={result['active_created']}, "
        f"active_existing={result['active_existing']}, "
        f"vip_granted={result['vip_granted']}, "
        f"vip_already_present={result['vip_already_present']}, "
        f"inactive_stored={result['inactive_stored']}, "
        f"inactive_existing_active={result['inactive_existing_active']}, "
        f"inactive_already_listed={result['inactive_already_listed']}"
    )


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    source = parser.add_mutually_exclusive_group(required=True)
    source.add_argument("--manifest", type=Path, help="Existing ID-only migration manifest")
    source.add_argument("--contacts-xlsx", type=Path, help="Legacy contacts XLSX")
    parser.add_argument(
        "--recipients-xlsx",
        type=Path,
        help="Verified legacy delivery XLSX; required with --contacts-xlsx",
    )
    parser.add_argument(
        "--write-manifest", type=Path, help="Write an ID-only manifest after XLSX validation"
    )
    parser.add_argument("--runtime-dir", type=Path, help="Override MARANIUS_RUNTIME_DIR")
    parser.add_argument(
        "--stage-runtime-dir",
        type=Path,
        help="Atomically stage the manifest for consumption during bot startup",
    )
    parser.add_argument(
        "--apply",
        action="store_true",
        help="Write users.json and legacy_inactive_users.json (default is preview only)",
    )
    args = parser.parse_args()
    if args.apply and args.stage_runtime_dir:
        parser.error("--apply and --stage-runtime-dir cannot be used together")

    if args.contacts_xlsx:
        if not args.recipients_xlsx or not args.write_manifest:
            parser.error("--contacts-xlsx requires --recipients-xlsx and --write-manifest")
        manifest = build_manifest(args.contacts_xlsx, args.recipients_xlsx)
        args.write_manifest.parent.mkdir(parents=True, exist_ok=True)
        args.write_manifest.write_text(
            json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
        )
        print(f"Wrote validated ID-only manifest: {args.write_manifest}")
    else:
        if args.recipients_xlsx or args.write_manifest:
            parser.error("--recipients-xlsx and --write-manifest are used only with --contacts-xlsx")
        manifest = load_manifest(args.manifest)

    if args.stage_runtime_dir:
        pending_path = args.stage_runtime_dir / "legacy_migration_pending.json"
        if pending_path.exists():
            raise ValueError(f"Pending migration already exists: {pending_path}")
        save_json(pending_path, manifest, trailing_newline=True)
        print(f"Staged pending migration: {pending_path}")
        return 0

    runtime_dir = args.runtime_dir or Path(os.getenv("MARANIUS_RUNTIME_DIR") or PROJECT_DIR)
    users_path = runtime_dir / "users.json"
    inactive_path = runtime_dir / "legacy_inactive_users.json"
    users = load_json(users_path, {})
    inactive_store = legacy_contacts.load_inactive_store(inactive_path)
    result = legacy_contacts.apply_legacy_migration(
        users,
        inactive_store,
        active_user_ids=manifest["active_user_ids"],
        vip_user_ids=manifest["vip_user_ids"],
        inactive_user_ids=manifest["inactive_user_ids"],
        source=str(manifest["source"]),
    )
    _print_summary(manifest=manifest, result=result, applied=args.apply)
    if not args.apply:
        print("No runtime data was changed. Re-run with --apply after checking this preview.")
        return 0

    # Offline escape hatch: use only while the bot is stopped.  The normal
    # production path is --stage-runtime-dir followed by the verified deploy.
    # Saving the inactive queue first makes a failed second write safe to rerun:
    # it can create a few redundant queued IDs, but cannot lose a returning user.
    save_json(inactive_path, inactive_store, trailing_newline=True)
    save_json(users_path, users, trailing_newline=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
